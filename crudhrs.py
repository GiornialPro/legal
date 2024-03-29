import os
from flask import request
import pandas as pd
from openpyxl import load_workbook, Workbook
import xlrd


class Sistema:

    def __init__(self, sourc):
        self.sourc = sourc

    def origina(self):
        
        loadb = load_workbook(self.sourc)
        booktotal = {'GERAL':''}

        for i in loadb.sheetnames:

            book = pd.read_excel(self.sourc, sheet_name='unidade')
            book_dic = book.to_dict('list')

            booktotal['GERAL'] = book_dic

        return booktotal

    def addciona(self):
        
        book_to_add = load_workbook(self.sourc)
        opn = pd.read_excel(self.sourc, sheet_name='unidade')
        opn_o = opn.to_dict('list')
        
        user = request.form['user']
        senha = request.form['senha']
        magic = request.form['magicword']

        # if request.method == 'POST':

        print(f'{magic}'.strip() == 'registrar', '##################')

        if f'{magic}'.strip() == 'registrar':
            pass
            
            book_to_add.create_sheet('Membros')


            book_to_add['unidade'][xlrd.cellname( 0, 0 )] = 'NOME'
            book_to_add['unidade'][xlrd.cellname( 0, 1 )] = 'CHAVE'
            book_to_add['unidade'][xlrd.cellname( 1, 0 )] = user
            book_to_add['unidade'][xlrd.cellname( 1, 1 )] = senha

        book_to_add.save(self.sourc)

        return senha, user, magic, f'new: {opn_o}', f'Folhas: {book_to_add.sheetnames}'
